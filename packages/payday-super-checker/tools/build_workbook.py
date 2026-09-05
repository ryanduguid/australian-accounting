"""Build workbooks/payday-super-checker.xlsx, the accountant-facing Excel front end.

The workbook restates the checker's rules in ordinary worksheet formulas so an
accountant with desktop Excel and no Python can paste the canonical contributions
register, set the as-at date and read the deadline, verdict and experimental SG
charge figures per line. It is pinned to the engine by ``tests/test_workbook.py``,
which reads the cached values desktop Excel wrote and compares them with ``assess``
on the shipped sample.

Run from the package directory on a machine with desktop Excel:

    uv run --locked --extra dev python tools/build_workbook.py

openpyxl writes the formulas; a single desktop Excel pass over COM recalculates the
workbook so the shipped file carries real cached values, then the build-machine path
Excel stamps into xl/workbook.xml is stripped.
"""

from __future__ import annotations

import csv
import json
import re
import subprocess
import sys
import zipfile
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import Table, TableColumn, TableFormula, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
from paydaysuper import __version__  # noqa: E402
from paydaysuper.assess import TRANSITION_END  # noqa: E402
from paydaysuper.csv_io import FALSE_WORDS, TRUE_WORDS  # noqa: E402
from paydaysuper.deadlines import REGIME_START  # noqa: E402
from paydaysuper.rates import days_in_year  # noqa: E402

OUT = ROOT / "workbooks" / "payday-super-checker.xlsx"
SAMPLE = ROOT / "examples" / "sample_payrun.csv"
DATA = ROOT / "paydaysuper" / "data"
DEFAULT_AS_AT = date(2026, 8, 10)
ESTIMATE_UNTIL = date(2030, 12, 31)

PURPLE, LAVENDER, GREY = "5C2D91", "F3F1F6", "E2E0DF"
HEAD = dict(font=Font(bold=True, color="FFFFFF"), fill=PatternFill("solid", fgColor=PURPLE))
INPUT = dict(font=Font(color="1F4E9E"), fill=PatternFill("solid", fgColor=GREY))
CALC = dict(fill=PatternFill("solid", fgColor=LAVENDER))
TEXT = "@"
MONEY = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
DATE = "yyyy-mm-dd"

INPUT_COLUMNS = [
    "employee_id", "payment_date", "sg_amount", "remitted_date", "remitted_amount",
    "matched_amount", "fund_received_date", "first_contribution_to_fund", "out_of_cycle",
    "next_standard_payday", "defined_benefit",
]
DATE_COLUMNS = ["payment_date", "remitted_date", "fund_received_date", "next_standard_payday"]
AMOUNT_COLUMNS = ["sg_amount", "remitted_amount", "matched_amount"]
BOOL_COLUMNS = ["first_contribution_to_fund", "out_of_cycle", "defined_benefit"]

AS_AT = "Summary!$B$2"
ASSESS = "Summary!$B$3"
TRANSITION_OK = "Summary!$B$4"
REMIT_ONLY_OK = "Summary!$B$5"
COVERAGE = "Summary!$B$6"
GIC_LAST = "Summary!$B$7"


def style(cell, **kw):
    for k, v in kw.items():
        setattr(cell, k, v)
    return cell


def header(ws, row, titles):
    for col, title in enumerate(titles, 1):
        style(ws.cell(row=row, column=col, value=title), **HEAD)


def add_table(ws, name, ref, formulas=None, array_columns=()):
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
    first_col, _, last_col, _ = range_boundaries(ref)
    columns = []
    for i, col in enumerate(range(first_col, last_col + 1), 1):
        title = ws.cell(row=1, column=col).value
        formula = (formulas or {}).get(title)
        columns.append(TableColumn(
            id=i, name=title,
            calculatedColumnFormula=None if formula is None else TableFormula(
                attr_text=formula, array=True if title in array_columns else None),
        ))
    table.tableColumns = columns
    ws.add_table(table)


def T(column):
    return f"tblLines[[#This Row],[{column}]]"


def excel_date(d):
    return f"DATE({d.year},{d.month},{d.day})"


def read_sample():
    with SAMPLE.open(newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def read_holidays():
    doc = json.loads((DATA / "business_days.json").read_text(encoding="utf-8"))
    rows = [
        (date.fromisoformat(h["date"]), h["name"], ", ".join(h["jurisdictions"]))
        for h in doc["non_business_days"]
        if not h["provisional"] and h["date"] >= "2026-01-01"
    ]
    provisional = [
        (date.fromisoformat(h["date"]), h["name"]) for h in doc["non_business_days"] if h["provisional"]
    ]
    return rows, provisional, date.fromisoformat(doc["verified_until"])


def read_gic():
    doc = json.loads((DATA / "gic_rates.json").read_text(encoding="utf-8"))
    rows = []
    for q in doc["quarters"]:
        start, end = date.fromisoformat(q["from"]), date.fromisoformat(q["to"])
        rows.append((start, end, float(q["annual_pct"]), days_in_year(start), "known", q["seen"]))
    last_end = rows[-1][1]
    last_pct = rows[-1][2]
    # Carry the last known rate forward in calendar-year segments, as daily_rate does
    # past the table, so the formula stays one SUMPRODUCT; staleness is flagged.
    start = last_end + timedelta(days=1)
    while start <= ESTIMATE_UNTIL:
        end = min(date(start.year, 12, 31), ESTIMATE_UNTIL)
        rows.append((start, end, last_pct, days_in_year(start), "estimate", ""))
        start = end + timedelta(days=1)
    return rows, last_end


def bool_expr(column):
    v = f"LOWER(TRIM({T(column)}&\"\"))"
    t = ",".join(f'"{x}"' for x in sorted(TRUE_WORDS))
    f = ",".join(f'"{x}"' for x in sorted(FALSE_WORDS))
    return f'IF(OR({v}={{{t}}}),1,IF(OR({v}={{{f}}}),0,""))'


def bd(start, n):
    return f"_xlfn.WORKDAY.INTL({start},{n},1,tblHolidays[date])"


# Branch codes: verdict, unassessable outer outcomes, whether the on-time receipt
# credit applies, whether the receipt is a stale pre-payment, whether exposed.
CODES = {
    "DB": ("SKIPPED", "", 0, 0, 0),
    "NIL": ("UNKNOWN", "", 0, 0, 0),
    "A1": ("ON_TIME", "", 1, 0, 0),
    "A2": ("UNKNOWN", "", 1, 0, 0),
    "A3": ("UNKNOWN", "UNPAID or NOT_YET_DUE", 1, 0, 0),
    "A4": ("UNPAID", "", 1, 0, 1),
    "S1": ("UNKNOWN", "", 0, 1, 0),
    "S2": ("UNKNOWN", "LATE or NOT_YET_DUE", 0, 1, 0),
    "S3": ("LATE", "", 0, 1, 1),
    "B1": ("UNKNOWN", "LATE or ON_TIME", 0, 0, 0),
    "B2": ("UNKNOWN", "LATE or NOT_YET_DUE", 0, 0, 0),
    "B3": ("UNKNOWN", "LATE or UNPAID", 0, 0, 0),
    "C1": ("UNKNOWN", "LATE or ON_TIME", 0, 0, 0),
    "C2": ("UNKNOWN", "LATE or NOT_YET_DUE", 0, 0, 0),
    "D1": ("ON_TIME", "", 1, 0, 0),
    "D2": ("UNKNOWN", "UNPAID or NOT_YET_DUE", 1, 0, 0),
    "D3": ("UNPAID", "", 1, 0, 1),
    "D4": ("UNKNOWN", "", 1, 0, 0),
    "E": ("LATE", "", 0, 0, 1),
    "R1": ("UNKNOWN", "LATE or AT_RISK", 0, 0, 0),
    "R2": ("UNKNOWN", "LATE or AT_RISK", 0, 0, 0),
    "R3": ("AT_RISK", "", 0, 0, 0),
    "R4": ("LATE", "", 0, 0, 1),
    "N1": ("UNKNOWN", "UNPAID or NOT_YET_DUE", 0, 0, 0),
    "N2": ("UNPAID", "", 0, 0, 1),
    "N3": ("UNKNOWN", "", 0, 0, 0),
}


def formulas():
    pay, sg = T("payment_date"), T("sg_amount")
    rem, rem_amt, matched = T("remitted_date"), T("remitted_amount"), T("matched_amount")
    rec, nxt = T("fund_received_date"), T("next_standard_payday")
    db, ooc, ftf = T("Flag_db"), T("Flag_ooc"), T("Flag_ftf")
    due, settled, remit = T("Final_due"), T("Settled"), T("Remit")
    poss = T("Possible_item4")
    horizon = f"({due}>{COVERAGE})"
    unc = f"({poss}<>\"\")"
    later_gate = f"OR({horizon},AND({unc},{AS_AT}<={poss}))"
    covers = f"({T('Receipt_credit')}>={sg})"
    nec_from = f"({due}+1)"
    nec_end = T("NEC_end")
    # Elementwise min and max via ABS: MIN and MAX aggregate a whole column inside
    # SUMPRODUCT, so the per-quarter overlap has to be written arithmetically.
    seg_end = f"(({nec_end}+tblGic[to]-ABS({nec_end}-tblGic[to]))/2)"
    seg_start = f"(({nec_from}+tblGic[from]+ABS({nec_from}-tblGic[from]))/2)"
    seg_days = f"({seg_end}-{seg_start}+1)"
    segments = f"(({seg_days}>0)*{seg_days})"
    v = T("Verdict")
    exposed = f"({T('Exposed')}=1)"
    date_bad = ",".join(f"AND({T(c)}<>\"\",NOT(ISNUMBER({T(c)})))" for c in DATE_COLUMNS)
    amount_bad = ",".join(
        f"AND({T(c)}<>\"\",OR(NOT(ISNUMBER({T(c)})),{T(c)}<0))" for c in AMOUNT_COLUMNS
    )
    bool_bad = ",".join(f'{T("Flag_" + s)}=""' for s in ("ftf", "ooc", "db"))
    return {
        "Flag_db": "=" + bool_expr("defined_benefit"),
        "Flag_ooc": "=" + bool_expr("out_of_cycle"),
        "Flag_ftf": "=" + bool_expr("first_contribution_to_fund"),
        "Row_problem": (
            f'=TRIM(IF(TRIM({T("employee_id")}&"")="","employee_id ","")'
            '&IF(OR(' + date_bad + '),"date ","")'
            f'&IF(NOT(ISNUMBER({pay})),"payment_date ","")'
            f'&IF(AND(ISNUMBER({pay}),{pay}<{excel_date(REGIME_START)}),"pre-1-Jul-2026 ","")'
            '&IF(OR(' + amount_bad + '),"amount ","")'
            f'&IF(NOT(ISNUMBER({sg})),"sg_amount ","")'
            '&IF(OR(' + bool_bad + '),"yes/no ","")'
            f'&IF(AND(ISNUMBER({rem_amt}),ISNUMBER({sg}),{rem_amt}>{sg}),"remitted_amount>sg ","")'
            f'&IF(AND(ISNUMBER({rem_amt}),NOT(ISNUMBER({rem}))),"remitted_amount-needs-date ","")'
            f'&IF(AND(ISNUMBER({matched}),ISNUMBER({sg}),{matched}>{sg}),"matched_amount>sg ","")'
            f'&IF(AND(ISNUMBER({matched}),ISNUMBER({rem_amt}),{rem_amt}>{matched}),"remitted_amount>matched ","")'
            f'&IF(AND(ISNUMBER({matched}),ISNUMBER({sg}),{matched}<{sg},ISNUMBER({rem}),NOT(ISNUMBER({rem_amt}))),"partial-match-needs-remitted_amount ","")'
            f'&IF(AND(ISNUMBER({rec}),ISNUMBER({rem}),{rec}<{rem}),"receipt-before-remittance ","")'
            f'&IF(AND({ooc}=1,NOT(ISNUMBER({nxt}))),"out_of_cycle-needs-next_standard_payday ","")'
            f'&IF(AND({ooc}=1,ISNUMBER({nxt}),ISNUMBER({pay}),{nxt}<={pay}),"next_standard_payday-not-after-payday ",""))'
        ),
        "Guard": "=IF(OR(" + ",".join(f"_xlfn.ISFORMULA({T(c)})" for c in INPUT_COLUMNS) + "),1,0)",
        "Cap": f'=IF(ISNUMBER({matched}),{matched},IF(ISNUMBER({rem_amt}),{rem_amt},{sg}))',
        "Due_ooc": f'=IF(AND({ooc}=1,ISNUMBER({nxt})),{bd(nxt, 7)},"")',
        "Due_ftf": f'=IF({ftf}=1,{bd(pay, 20)},"")',
        # 0, not "", where there is no deadline: the item 4 lookups multiply this column
        # inside SUMPRODUCT and text would raise #VALUE! even when multiplied by zero.
        "Own_due": (
            f'=IF(OR({db}=1,NOT(ISNUMBER({pay}))),0,IF(OR({ooc}=1,{ftf}=1),'
            f'MAX({T("Due_ooc")},{T("Due_ftf")}),{bd(pay, 7)}))'
        ),
        "Own_pathway": (
            f'=IF({db}=1,"SKIP_DB",IF({ooc}=1,IF(AND({ftf}=1,{T("Due_ftf")}>{T("Due_ooc")}),'
            f'"EXTENDED_20BD","OUT_OF_CYCLE"),IF({ftf}=1,"EXTENDED_20BD","USUAL_7BD")))'
        ),
        "Earliest_prepay": f'=IF(ISNUMBER({pay}),_xlfn.EDATE({pay}-1,-12)+1,"")',
        "Settled": f'=IF(AND(ISNUMBER({rec}),{rec}<={AS_AT}),{rec},"")',
        "Remit": f'=IF(AND(ISNUMBER({rem}),{rem}<={AS_AT}),{rem},"")',
        "Receipt_credit": f'=IF({settled}<>"",MIN(ROUND({T("Cap")},2),ROUND({sg},2)),0)',
        "Credit": (
            f'=IF(ISNUMBER({rem_amt}),IF({remit}<>"",MIN(ROUND({rem_amt},2),ROUND({sg},2)),0),'
            f'IF({remit}<>"",ROUND({sg},2),0))'
        ),
        # Evidence against the row's own deadline. Aligned or possible deadlines only
        # ever propagate values already present, so the sweep needs no recursion.
        "Evidence_own": (
            f'=IF(OR({T("Own_due")}=0,{sg}<=0,{db}=1,{T("Cap")}<=0),"impossible",'
            f'IF(ISNUMBER({rec}),IF({rec}<{T("Earliest_prepay")},"impossible",'
            f'IF({rec}<={AS_AT},IF(OR({rec}<{pay},{rec}<={T("Own_due")}),"confirmed",'
            f'IF({rec}>{T("Own_due")},"impossible","possible")),'
            f'IF({rec}<={T("Own_due")},"possible","impossible"))),'
            f'IF(AND(ISNUMBER({rem}),{rem}>{T("Own_due")}),"impossible","possible")))'
        ),
        # EXACT keeps employee ids case-sensitive, as the engine groups them; MAXIFS
        # would fold E123 and e123 together and could extend a deadline it should not.
        "Confirmed_latest": (
            f'=IF(OR({db}=1,{T("Own_due")}=0),"",SUMPRODUCT(MAX(EXACT(tblLines[employee_id],{T("employee_id")})'
            f'*(tblLines[payment_date]<{pay})*(tblLines[Evidence_own]="confirmed")*tblLines[Own_due])))'
        ),
        "Possible_latest": (
            f'=IF(OR({db}=1,{T("Own_due")}=0),"",SUMPRODUCT(MAX(EXACT(tblLines[employee_id],{T("employee_id")})'
            f'*(tblLines[payment_date]<{pay})*(tblLines[Evidence_own]<>"impossible")*tblLines[Own_due])))'
        ),
        "Final_due": (
            f'=IF({T("Own_due")}=0,"",IF(AND(ISNUMBER({T("Confirmed_latest")}),'
            f'{T("Confirmed_latest")}>{T("Own_due")}),{T("Confirmed_latest")},{T("Own_due")}))'
        ),
        "Pathway": (
            f'=IF({T("Own_due")}=0,{T("Own_pathway")},IF(AND(ISNUMBER({T("Confirmed_latest")}),'
            f'{T("Confirmed_latest")}>{T("Own_due")}),"ITEM4_ALIGNED",{T("Own_pathway")}))'
        ),
        "Case_variant": (
            f'=IF(COUNTIF(tblLines[employee_id],{T("employee_id")})>SUMPRODUCT(--EXACT(tblLines[employee_id],'
            f'{T("employee_id")})),1,0)'
        ),
        "Sample_row": "=IF(AND(ISNUMBER(MATCH(" + T("employee_id") + ",{SAMPLE_IDS},0)),ISNUMBER(MATCH("
                      + pay + ",{SAMPLE_DATES},0))),1,0)",
        "Possible_item4": (
            f'=IF({due}="","",IF(AND(ISNUMBER({T("Possible_latest")}),{T("Possible_latest")}>{due}),'
            f'{T("Possible_latest")},""))'
        ),
        "Past_horizon": f'=IF({due}="","",IF({horizon},1,0))',
        "Branch": (
            f'=IF({db}=1,"DB",IF({due}="","",IF({sg}<=0,"NIL",'
            f'IF({settled}<>"",'
            # received branch
            f'IF({settled}<{pay},'
            f'IF({settled}>={T("Earliest_prepay")},'
            f'IF({covers},"A1",IF({due}>={AS_AT},"A2",IF({later_gate},"A3","A4"))),'
            f'IF({due}>={AS_AT},"S1",IF({later_gate},"S2","S3"))),'
            f'IF(AND({unc},{due}<{settled},{settled}<={poss}),'
            f'IF({covers},"B1",IF({AS_AT}<={poss},"B2","B3")),'
            f'IF(AND({horizon},{settled}>{due}),IF({covers},"C1","C2"),'
            f'IF({settled}<={due},IF({covers},"D1",IF({later_gate},"D2",IF({due}<{AS_AT},"D3","D4"))),"E")))),'
            # without receipt
            f'IF(AND({remit}<>"",{sg}-{T("Credit")}<=0),'
            f'IF(AND({unc},{due}<{remit},{remit}<={poss}),"R1",'
            f'IF(AND({horizon},{remit}>{due}),"R2",IF({remit}<={due},"R3","R4"))),'
            f'IF({due}<{AS_AT},IF({later_gate},"N1","N2"),"N3"))))))'
        ),
        "Verdict": f'=IF({T("Branch")}="","",INDEX(tblCodes[verdict],MATCH({T("Branch")},tblCodes[code],0)))',
        # T() turns the 0 that INDEX returns for an empty Codes cell back into "".
        "Unassessable_between": (
            f'=IF({T("Branch")}="","",T(INDEX(tblCodes[unassessable],MATCH({T("Branch")},tblCodes[code],0))))'
        ),
        "Exposed": f'=IF({T("Branch")}="",0,INDEX(tblCodes[exposed],MATCH({T("Branch")},tblCodes[code],0)))',
        "Stale_prepay": f'=IF({T("Branch")}="",0,INDEX(tblCodes[stale],MATCH({T("Branch")},tblCodes[code],0)))',
        "OTRC": (
            f'=IF({T("Branch")}="",0,INDEX(tblCodes[otrc],MATCH({T("Branch")},tblCodes[code],0))*{T("Receipt_credit")})'
        ),
        "Base_shortfall": f'=IF({exposed},MAX(ROUND({sg},2)-{T("OTRC")},0),"")',
        "Offset_s18D": (
            f'=IF({exposed},IF(AND({settled}<>"",{T("Stale_prepay")}=0,{settled}>{due},'
            f'OR({ASSESS}="",{settled}<{ASSESS})),MIN({T("Receipt_credit")},{T("Base_shortfall")}),0),"")'
        ),
        "Final_shortfall": f'=IF({exposed},MAX({T("Base_shortfall")}-{T("Offset_s18D")},0),"")',
        "Lateness_basis": (
            f'=IF({exposed},IF(AND({settled}<>"",{T("Stale_prepay")}=0,{covers}),"fund receipt",'
            f'IF(AND({settled}<>"",{T("Stale_prepay")}=0),"as-at date (shortfall remains after part receipt)",'
            f'IF(AND(OR({settled}<>"",{remit}<>""),{T("Stale_prepay")}=0),"as-at date (no fund receipt recorded)",'
            f'"as-at date (nothing applied to this payday)"))),"")'
        ),
        "Outstanding_to": (
            f'=IF({exposed},IF(AND({settled}<>"",{T("Stale_prepay")}=0,{covers}),{settled},{AS_AT}),"")'
        ),
        "NEC_end": (
            f'=IF({exposed},MIN(IF({ASSESS}="",{excel_date(ESTIMATE_UNTIL)},{ASSESS}-1),'
            f'IF(AND({settled}<>"",{T("Stale_prepay")}=0,{covers},{T("Final_shortfall")}=0),{settled},{AS_AT})),"")'
        ),
        "Days_late": (
            f'=IF({exposed},IF({T("Past_horizon")}=1,"",MAX({T("Outstanding_to")}-{due},0)),"")'
        ),
        "NEC": (
            f'=IF({exposed},IF({nec_end}>{due},{T("Base_shortfall")}*(EXP(SUMPRODUCT({segments}'
            f'*LN(1+tblGic[annual_pct]/100/tblGic[divisor])))-1),0),"")'
        ),
        "GIC_estimated": f'=IF({exposed},IF({nec_end}>{GIC_LAST},1,0),"")',
        "Shortfall_r": f'=IF({exposed},ROUND({T("Final_shortfall")},2),"")',
        "NEC_r": f'=IF({exposed},ROUND({T("NEC")},2),"")',
        "Uplift_best": f'=IF({exposed},0,"")',
        "Uplift_worst": f'=IF({exposed},ROUND(0.6*({T("Final_shortfall")}+{T("NEC")}),2),"")',
        "SGC_low": f'=IF({exposed},{T("Shortfall_r")}+{T("NEC_r")}+{T("Uplift_best")},"")',
        "SGC_high": f'=IF({exposed},{T("Shortfall_r")}+{T("NEC_r")}+{T("Uplift_worst")},"")',
        "Transition_row": (
            f'=IF(AND({db}<>1,ISNUMBER({sg}),{sg}>0,{T("Cap")}>0,OR(ISNUMBER({rec}),ISNUMBER({rem})),'
            f'IF(ISNUMBER({rec}),{rec},{rem})<={excel_date(TRANSITION_END)}),1,0)'
        ),
        "Receipt_established": f'=IF({settled}<>"",1,0)',
        "Assessable": f'=IF(AND({v}<>"",{v}<>"SKIPPED",ISNUMBER({sg}),{sg}>0),1,0)',
        "Duplicate": (
            f'=IF(COUNTIFS(tblLines[employee_id],{T("employee_id")},tblLines[payment_date],{pay},'
            f'tblLines[sg_amount],{sg},tblLines[remitted_date],{rem}&"",tblLines[fund_received_date],{rec}&"")>1,1,0)'
        ),
    }


CALC_ORDER = [
    "Flag_db", "Flag_ooc", "Flag_ftf", "Row_problem", "Guard", "Cap", "Due_ooc", "Due_ftf",
    "Own_due", "Own_pathway", "Earliest_prepay", "Settled", "Remit", "Receipt_credit", "Credit",
    "Evidence_own", "Confirmed_latest", "Possible_latest", "Final_due", "Pathway",
    "Possible_item4", "Past_horizon", "Branch", "Verdict", "Unassessable_between", "Exposed",
    "Stale_prepay", "OTRC", "Base_shortfall", "Offset_s18D", "Final_shortfall", "Lateness_basis",
    "Outstanding_to", "NEC_end", "Days_late", "NEC", "GIC_estimated", "Shortfall_r", "NEC_r",
    "Uplift_best", "Uplift_worst", "SGC_low", "SGC_high", "Transition_row",
    "Receipt_established", "Assessable", "Duplicate", "Case_variant", "Sample_row",
]
ARRAY_CALCS = {"Confirmed_latest", "Possible_latest", "Case_variant"}
DATE_CALCS = {"Due_ooc", "Due_ftf", "Own_due", "Earliest_prepay", "Settled", "Remit",
              "Confirmed_latest", "Possible_latest", "Final_due", "Possible_item4",
              "Outstanding_to", "NEC_end"}
MONEY_CALCS = {"Cap", "Receipt_credit", "Credit", "OTRC", "Base_shortfall", "Offset_s18D",
               "Final_shortfall", "NEC", "Shortfall_r", "NEC_r", "Uplift_best", "Uplift_worst",
               "SGC_low", "SGC_high"}


def build() -> None:
    rows = read_sample()
    holidays, provisional, verified_until = read_holidays()
    gic, gic_last = read_gic()
    calc = formulas()
    epoch = date(1899, 12, 30)
    sample_ids = ",".join(f'"{r["employee_id"]}"' for r in rows)
    sample_dates = ",".join(str((date.fromisoformat(r["payment_date"]) - epoch).days) for r in rows)
    calc["Sample_row"] = (calc["Sample_row"].replace("{SAMPLE_IDS}", "{" + sample_ids + "}")
                          .replace("{SAMPLE_DATES}", "{" + sample_dates + "}"))
    assert set(calc) == set(CALC_ORDER), set(calc) ^ set(CALC_ORDER)
    wb = Workbook()

    # 1. Start Here
    ws = wb.active
    ws.title = "Start Here"
    style(ws["A1"], value="Payday Super review: s 18C deadlines and the SG charge estimate",
          font=Font(bold=True, size=14, color=PURPLE))
    steps = [
        "1. On Register, paste one contribution line per row over the example rows, in the "
        "eleven canonical columns the command line tool reads and the importer writes. Dates are "
        "real Excel dates; amounts are dollars; yes/no columns take yes, no or blank (no).",
        "2. On Summary, set the as-at date (notional earnings on unpaid lines run to it), an "
        "assessment date if the ATO has assessed, and the two confirmations the command line "
        "tool asks for: LCR 2026/1 transition allocation and remittance-only review.",
        "3. Read the calculated columns on Register: pathway, final deadline, verdict "
        "(ON_TIME, AT_RISK, LATE, UNPAID, UNKNOWN, SKIPPED), the conservative outer outcomes "
        "where the facts cannot decide, days late, final shortfall after s 18D, notional "
        "earnings, the uplift range and the experimental SG charge estimate range.",
        "4. Review Checks: BLOCKED means the register cannot be read or a required "
        "confirmation is missing. REVIEW means a line is exposed, a line is undecided, or the "
        "file cannot produce ON_TIME and remittance-only review is not confirmed. PASS mirrors "
        "the command line tool's exit code 0.",
        "5. The statutory test is receipt by the fund. A remittance date alone gives AT_RISK, "
        "never ON_TIME. Fill fund_received_date from the clearing house or fund before treating "
        "a verdict as final.",
        "6. Experimental review aid. Not a compliance determination. The ATO makes the "
        "assessment; choice loading, the maximum contributions base and post-assessment "
        "charges are not estimated.",
    ]
    for i, text in enumerate(steps, 3):
        ws.cell(row=i, column=1, value=text).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110
    style(ws["A10"], value="Overall status", font=Font(bold=True))
    style(ws["A11"], value="='Review Checks'!B16", font=Font(bold=True, size=14), **CALC)
    ws["A13"] = (f"payday-super-checker workbook, engine {__version__}. Macro-free. "
                 "Not advice; see DISCLAIMER.md in the repository.")

    # 2. Register
    ws = wb.create_sheet("Register")
    titles = INPUT_COLUMNS + CALC_ORDER
    header(ws, 1, titles)
    for r, row in enumerate(rows, 2):
        for c, column in enumerate(INPUT_COLUMNS, 1):
            raw = row.get(column, "")
            if column in DATE_COLUMNS:
                value = date.fromisoformat(raw) if raw else None
            elif column in AMOUNT_COLUMNS:
                value = float(raw) if raw else None
            else:
                value = raw or None
            cell = style(ws.cell(row=r, column=c, value=value), **INPUT)
            if column in DATE_COLUMNS:
                cell.number_format = DATE
            elif column in AMOUNT_COLUMNS:
                cell.number_format = MONEY
            else:
                cell.number_format = TEXT
        for c, column in enumerate(CALC_ORDER, len(INPUT_COLUMNS) + 1):
            cell = style(ws.cell(row=r, column=c), **CALC)
            # Array formulas: a same-table column inside EXACT would otherwise be
            # rewritten to a this-row reference on load and match every employee.
            if column in ARRAY_CALCS:
                cell.value = ArrayFormula(cell.coordinate, calc[column])
            else:
                cell.value = calc[column]
            if column in DATE_CALCS:
                cell.number_format = DATE
            elif column in MONEY_CALCS:
                cell.number_format = MONEY
    last_row = len(rows) + 1
    add_table(ws, "tblLines", f"A1:{get_column_letter(len(titles))}{last_row}",
              {k: v[1:] for k, v in calc.items()}, ARRAY_CALCS)
    for column in BOOL_COLUMNS:
        col = get_column_letter(INPUT_COLUMNS.index(column) + 1)
        dv = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
        dv.add(f"{col}2:{col}{last_row + 1000}")
        ws.add_data_validation(dv)
    for c, column in enumerate(titles, 1):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.column_dimensions[get_column_letter(titles.index("Row_problem") + 1)].width = 30
    ws.column_dimensions[get_column_letter(titles.index("Lateness_basis") + 1)].width = 34
    ws.column_dimensions[get_column_letter(titles.index("Unassessable_between") + 1)].width = 22
    ws.freeze_panes = "B2"

    # 3. Codes (verdict lookup)
    ws = wb.create_sheet("Codes")
    header(ws, 1, ["code", "verdict", "unassessable", "otrc", "stale", "exposed"])
    for r, (code, values) in enumerate(CODES.items(), 2):
        ws.cell(row=r, column=1, value=code)
        for c, value in enumerate(values, 2):
            ws.cell(row=r, column=c, value=value)
    add_table(ws, "tblCodes", f"A1:F{len(CODES) + 1}")
    ws.column_dimensions["C"].width = 24
    ws.protection.sheet = True

    # 4. Holidays
    ws = wb.create_sheet("Holidays")
    header(ws, 1, ["date", "name", "jurisdictions"])
    for r, (d, name, juris) in enumerate(holidays, 2):
        style(ws.cell(row=r, column=1, value=d), number_format=DATE, **INPUT)
        style(ws.cell(row=r, column=2, value=name), **INPUT)
        style(ws.cell(row=r, column=3, value=juris), **INPUT)
    add_table(ws, "tblHolidays", f"A1:C{len(holidays) + 1}")
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 40
    note_row = len(holidays) + 4
    ws.cell(row=note_row, column=1, value=(
        "Whole-of-jurisdiction public holidays under SGAA s 6(1); weekends are computed. "
        f"Official sources checked through {verified_until.isoformat()}. To review a later "
        "deadline, add the official dates here and move the coverage date on Summary."))
    ws.cell(row=note_row + 1, column=1, value=(
        "Provisional dates not applied (business days until confirmed): "
        + "; ".join(f"{d.isoformat()} {n}" for d, n in provisional)))

    # 5. GIC
    ws = wb.create_sheet("GIC")
    header(ws, 1, ["from", "to", "annual_pct", "divisor", "basis", "seen"])
    for r, (start, end, pct, divisor, basis, seen) in enumerate(gic, 2):
        style(ws.cell(row=r, column=1, value=start), number_format=DATE, **INPUT)
        style(ws.cell(row=r, column=2, value=end), number_format=DATE, **INPUT)
        style(ws.cell(row=r, column=3, value=pct), **INPUT)
        style(ws.cell(row=r, column=4, value=divisor), **INPUT)
        style(ws.cell(row=r, column=5, value=basis), **INPUT)
        style(ws.cell(row=r, column=6, value=seen), **INPUT)
    add_table(ws, "tblGic", f"A1:F{len(gic) + 1}")
    ws.cell(row=len(gic) + 4, column=1, value=(
        "General interest charge, TAA 1953 s 8AAD: annual rate divided by the days in the "
        "calendar year. Rows marked estimate carry the last known rate forward, as the "
        "engine does, and the workbook flags any line that uses them. Update each quarter "
        "from the ATO GIC rates page and set basis to known."))
    ws.column_dimensions["F"].width = 14

    # 6. Summary
    ws = wb.create_sheet("Summary")
    style(ws["A1"], value="Inputs", **HEAD)
    style(ws["B1"], value="", **HEAD)
    inputs = [
        (2, "As-at date (notional earnings on unpaid lines run to it)", DEFAULT_AS_AT, DATE),
        (3, "ATO assessment date (blank if none has issued)", None, DATE),
        (4, "LCR 2026/1 transition allocation reconciled and confirmed (Y/N)", "N", None),
        (5, "Remittance-only review accepted (Y/N)", "N", None),
        (6, "Holiday coverage verified until", verified_until, DATE),
    ]
    for r, label, value, fmt in inputs:
        ws.cell(row=r, column=1, value=label)
        cell = style(ws.cell(row=r, column=2, value=value), protection=Protection(locked=False), **INPUT)
        if fmt:
            cell.number_format = fmt
    ws["A7"] = "GIC table last known quarter ends"
    style(ws["B7"], value='=_xlfn.MAXIFS(tblGic[to],tblGic[basis],"known")', number_format=DATE, **CALC)
    for ref in ("B4", "B5"):
        dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
        dv.add(ref)
        ws.add_data_validation(dv)
    for ref, blank_ok in (("B2", False), ("B3", True), ("B6", False)):
        dv = DataValidation(type="date", operator="greaterThan", formula1="1", allow_blank=blank_ok)
        dv.add(ref)
        ws.add_data_validation(dv)
    style(ws["A9"], value="Summary", **HEAD)
    style(ws["B9"], value="", **HEAD)
    summary = [
        ("ON_TIME", '=COUNTIF(tblLines[Verdict],"ON_TIME")'),
        ("AT_RISK", '=COUNTIF(tblLines[Verdict],"AT_RISK")'),
        ("LATE", '=COUNTIF(tblLines[Verdict],"LATE")'),
        ("UNPAID", '=COUNTIF(tblLines[Verdict],"UNPAID")'),
        ("UNKNOWN", '=COUNTIF(tblLines[Verdict],"UNKNOWN")'),
        ("SKIPPED", '=COUNTIF(tblLines[Verdict],"SKIPPED")'),
        ("Lines with exposure", "=SUM(tblLines[Exposed])"),
        ("Total final shortfall", "=ROUND(SUM(tblLines[Shortfall_r]),2)"),
        ("Total notional earnings", "=ROUND(SUM(tblLines[NEC_r]),2)"),
        ("Experimental SG charge estimate, low", "=ROUND(SUM(tblLines[SGC_low]),2)"),
        ("Experimental SG charge estimate, high", "=ROUND(SUM(tblLines[SGC_high]),2)"),
        ("Attention-driving UNKNOWN lines", '=SUMPRODUCT(--(tblLines[Unassessable_between]<>""))'),
        ("Remittance-only file (no assessable line has a fund receipt)",
         '=IF(AND(SUM(tblLines[Assessable])>0,SUMPRODUCT(tblLines[Assessable]*tblLines[Receipt_established])=0),"YES","NO")'),
    ]
    for r, (label, formula) in enumerate(summary, 10):
        ws.cell(row=r, column=1, value=label)
        cell = style(ws.cell(row=r, column=2, value=formula), **CALC)
        if "shortfall" in label or "earnings" in label or "estimate" in label:
            cell.number_format = MONEY
    ws["A24"] = "Overall status"
    style(ws["B24"], value="='Review Checks'!B16", font=Font(bold=True), **CALC)
    ws.column_dimensions["A"].width = 64
    ws.column_dimensions["B"].width = 18
    ws.protection.sheet = True

    # 7. Review Checks
    ws = wb.create_sheet("Review Checks")
    header(ws, 1, ["Check", "Result", "Count or value", "Example row (employee_id)"])

    def offender(cond):
        return f"INDEX(tblLines[employee_id],SUMPRODUCT(MAX({cond}*(ROW(tblLines[employee_id])-1))))"

    checks = [
        ("Every register value can be read", '=IF(C2=0,"PASS","BLOCKED")',
         '=SUMPRODUCT(--(tblLines[Row_problem]<>""))',
         '=IF(C2=0,"",' + offender('(tblLines[Row_problem]<>"")') + ")"),
        ("No formula in an input cell", '=IF(C3=0,"PASS","BLOCKED")', "=SUM(tblLines[Guard])",
         '=IF(C3=0,"",' + offender("(tblLines[Guard]=1)") + ")"),
        ("Contributions dated on or before 28 July 2026 are reconciled under LCR 2026/1",
         f'=IF(OR(C4=0,{TRANSITION_OK}="Y"),"PASS","BLOCKED")', "=SUM(tblLines[Transition_row])",
         '=IF(C4=0,"",' + offender("(tblLines[Transition_row]=1)") + ")"),
        ("No line is LATE or UNPAID", '=IF(C5=0,"PASS","REVIEW")', "=SUM(tblLines[Exposed])",
         '=IF(C5=0,"",' + offender("(tblLines[Exposed]=1)") + ")"),
        ("No line is left undecided by the deadline facts", '=IF(C6=0,"PASS","REVIEW")',
         '=SUMPRODUCT(--(tblLines[Unassessable_between]<>""))',
         '=IF(C6=0,"",' + offender('(tblLines[Unassessable_between]<>"")') + ")"),
        ("The file can produce ON_TIME, or remittance-only review is accepted",
         f'=IF(OR(Summary!B22="NO",{REMIT_ONLY_OK}="Y"),"PASS","REVIEW")', "=Summary!B22", None),
        ("Deadlines inside the holiday coverage (figures past it are a maximum)",
         '=IF(C8=0,"PASS","NOTE")', "=SUM(tblLines[Past_horizon])",
         '=IF(C8=0,"",' + offender("(tblLines[Past_horizon]=1)") + ")"),
        ("Notional earnings use only known GIC quarters", '=IF(C9=0,"PASS","NOTE")',
         "=SUM(tblLines[GIC_estimated])",
         '=IF(C9=0,"",' + offender("(tblLines[GIC_estimated]=1)") + ")"),
        ("No identical lines (a doubled export counts a payday twice)", '=IF(C10=0,"PASS","NOTE")',
         "=SUM(tblLines[Duplicate])",
         '=IF(C10=0,"",' + offender("(tblLines[Duplicate]=1)") + ")"),
        ("Lines assessed at a nil SG amount (nothing to assess)", '=IF(C11=0,"PASS","NOTE")',
         '=COUNTIF(tblLines[Branch],"NIL")', '=IF(C11=0,"",' + offender('(tblLines[Branch]="NIL")') + ")"),
        ("As-at, assessment and coverage dates on Summary are dates",
         f'=IF(AND(ISNUMBER({AS_AT}),OR({ASSESS}="",ISNUMBER({ASSESS})),ISNUMBER({COVERAGE})),"PASS","BLOCKED")',
         f'={AS_AT}', None),
        ("No fabricated example line from the shipped sample remains in the register",
         '=IF(C13=0,"PASS","REVIEW")', "=SUM(tblLines[Sample_row])",
         '=IF(C13=0,"",' + offender("(tblLines[Sample_row]=1)") + ")"),
        ("No employee ids differ only by capitalisation (treated as different people, not aligned)",
         '=IF(C14=0,"PASS","NOTE")', "=SUM(tblLines[Case_variant])",
         '=IF(C14=0,"",' + offender("(tblLines[Case_variant]=1)") + ")"),
    ]
    for r, (label, result, detail, example) in enumerate(checks, 2):
        ws.cell(row=r, column=1, value=label)
        style(ws.cell(row=r, column=2, value=result), **CALC)
        style(ws.cell(row=r, column=3, value=detail), **CALC)
        cell = style(ws.cell(row=r, column=4), **CALC)
        if example:
            cell.value = ArrayFormula(cell.coordinate, example)
    last = len(checks) + 1
    style(ws["A16"], value="Overall status", font=Font(bold=True))
    style(ws["B16"], value=(f'=IF(COUNTIF(B2:B{last},"BLOCKED")>0,"BLOCKED",'
                            f'IF(COUNTIF(B2:B{last},"REVIEW")>0,"REVIEW","PASS"))'),
          font=Font(bold=True), **CALC)
    style(ws["A18"], value="Assumptions the engine prints with every run", **HEAD)
    caveats = [
        "The statutory test is receipt by the fund with enough information to allocate it "
        "(SGAA s 18C(1)(c)). A remittance date is operational evidence only; clearing-house "
        "transit is the employer's risk.",
        "Item 4 (s 18C(2)) aligns a later payday only to an earlier eligible contribution the "
        "register evidences as received on time; a positive amount or a remittance alone gives "
        "a possible upper bound and an attention-driving UNKNOWN where it would change the verdict.",
        "Notional earnings compound daily at the general interest charge on the base shortfall "
        "from the day after the deadline while the final shortfall remains greater than nil "
        "(s 19A, LCR 2026/3). A part receipt does not slow the accrual.",
        "The uplift range runs from 0 per cent (clean history and a voluntary disclosure within "
        "30 days) to 60 per cent (prior history, no disclosure). The ATO, not this workbook, "
        "decides which reductions apply.",
        "Each exposure component is rounded to cents half up and the totals are built from the "
        "rounded parts. TAA 1953 s 16B applies a five-cent down-round only to the Commissioner's "
        "final assessed charge, which is not reproduced.",
        "Not estimated: choice loading, the maximum contributions base, the late payment "
        "penalty, interest after assessment, exceptional-circumstances determinations, fund deed "
        "and award obligations, and paydays before 1 July 2026.",
    ]
    for r, text in enumerate(caveats, 19):
        ws.cell(row=r, column=1, value=text).alignment = Alignment(wrap_text=True)
    fills = {"BLOCKED": "F8D7DA", "REVIEW": "FFF3CD", "PASS": "D4EDDA"}
    for target, ref in ((ws, "B2:B16"), (wb["Start Here"], "A11"), (wb["Summary"], "B24")):
        for word, colour in fills.items():
            target.conditional_formatting.add(
                ref, CellIsRule(operator="equal", formula=[f'"{word}"'],
                                fill=PatternFill("solid", fgColor=colour)))
    ws.column_dimensions["A"].width = 100
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 24
    ws.protection.sheet = True

    # 8. Sources & Version
    ws = wb.create_sheet("Sources & Version")
    header(ws, 1, ["Item", "Value"])
    items = [
        ("Workbook engine version", __version__),
        ("Desktop Excel build used to calculate cached values", ""),
        ("Law", "Superannuation Guarantee (Administration) Act 1992 as amended by the Treasury "
                "Laws Amendment (Payday Superannuation) Act 2025 (No. 57 of 2025) and the "
                "Superannuation Guarantee Charge Amendment Act 2025 (No. 58 of 2025); "
                "Payday Superannuation Regulations 2026 (F2026L00133); F2026L00784; "
                "LCR 2026/1, 2026/2 and 2026/3"),
        ("Primary-source review", "docs/primary-source-review-2026-08-15.md in the repository"),
        ("Holiday calendar", f"Official whole-of-jurisdiction sources checked through {verified_until.isoformat()}"),
        ("GIC rates", f"ATO general interest charge rates, known through {gic_last.isoformat()}; "
                      "later days carry the last known rate as an estimate"),
        ("No-install explainer", "https://duguid.com.au/tools/payday-super/"),
        ("Disclaimer", "Experimental review aid. Not a compliance determination, not an ATO "
                       "assessment, and not tax, legal or financial advice."),
    ]
    for r, (k, v) in enumerate(items, 2):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 110

    OUT.parent.mkdir(exist_ok=True)
    wb.save(OUT)


RECALC = r"""
$ErrorActionPreference = 'Stop'
$t = 0; while ((Get-Process EXCEL -ErrorAction SilentlyContinue) -and $t -lt 60) { Start-Sleep 1; $t++ }
if (Get-Process EXCEL -ErrorAction SilentlyContinue) {
  throw 'Excel is already running; refusing to share its COM server.'
}
$xl = New-Object -ComObject Excel.Application
$xl.Visible = $false; $xl.DisplayAlerts = $false; $xl.AutomationSecurity = 1
try {
  $wb = $xl.Workbooks.Open('%s')
  $sources = $wb.Worksheets.Item('Sources & Version')
  $sources.Range('B3').Value2 = 'Excel ' + $xl.Version + ' build ' + $xl.Build
  $xl.CalculateFullRebuild()
  $tries = 0
  while ($xl.CalculationState -ne 0 -and $tries -lt 600) { Start-Sleep -Milliseconds 100; $tries++ }
  $status = $wb.Worksheets.Item('Review Checks').Range('B16').Text
  $wb.Save()
  $wb.Close($false)
  Write-Output ('overall=' + $status)
} finally { $xl.Quit() }
"""


def recalc() -> None:
    script = RECALC % str(OUT).replace("'", "''")
    result = subprocess.run(["powershell", "-NoProfile", "-Command", script],
                            capture_output=True, text=True)
    if result.returncode:
        raise SystemExit(result.stderr or result.stdout)
    print(result.stdout.strip())
    tmp = OUT.with_suffix(".tmp")
    with zipfile.ZipFile(OUT) as src, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "xl/workbook.xml":
                data = re.sub(rb"<x15ac:absPath[^>]*/>", b"", data)
            dst.writestr(item, data)
    tmp.replace(OUT)


if __name__ == "__main__":
    build()
    if "--no-excel" not in sys.argv:
        recalc()
    print(OUT)
