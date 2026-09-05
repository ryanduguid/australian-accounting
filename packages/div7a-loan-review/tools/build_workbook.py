"""Build workbooks/div7a-loan-review.xlsx, the accountant-facing Excel front end.

The workbook restates the engine's rules in ordinary worksheet formulas so an
accountant with desktop Excel and no Python can paste a loan register, pick the year
of income and read the s 109N gate and the s 109E minimum yearly repayment per loan.
It is pinned to the engine by ``tests/test_workbook.py``, which reads the cached
values desktop Excel wrote and compares them with ``review_register`` on the mixed
sample.

Run from the package directory on a machine with desktop Excel:

    uv run --locked --extra dev python tools/build_workbook.py

openpyxl writes the formulas; a single desktop Excel pass over COM recalculates the
workbook so the shipped file carries real cached values, then the build-machine path
Excel stamps into xl/workbook.xml is stripped.
"""

from __future__ import annotations

import csv
import re
import subprocess
import sys
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import Table, TableColumn, TableFormula, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
from div7aloan import __version__  # noqa: E402
from div7aloan.facts import UNKNOWN_TOKENS, _FALSE, _TRUE  # noqa: E402
from div7aloan.register import FIRST_REVIEWABLE_YEAR  # noqa: E402

OUT = ROOT / "workbooks" / "div7a-loan-review.xlsx"
SAMPLE = ROOT / "examples" / "sample_loans_mixed.csv"
RATES = ROOT / "div7aloan" / "data" / "benchmark_rates.csv"
DEFAULT_YEAR = "2026-27"

PURPLE, LAVENDER, GREY = "5C2D91", "F3F1F6", "E2E0DF"
HEAD = dict(font=Font(bold=True, color="FFFFFF"), fill=PatternFill("solid", fgColor=PURPLE))
INPUT = dict(font=Font(color="1F4E9E"), fill=PatternFill("solid", fgColor=GREY))
CALC = dict(fill=PatternFill("solid", fgColor=LAVENDER))
TEXT = "@"
MONEY = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'

INPUT_COLUMNS = [
    "loan_id", "borrower_reference", "year_loan_made", "written_agreement",
    "terms_in_place_before_lodgment_day", "maximum_term_years",
    "secured_by_registered_mortgage_over_real_property", "security_coverage_at_first_made",
    "interest_rate_for_years_after_year_loan_made",
    "amalgamated_loan_unpaid_at_end_of_previous_year", "remaining_term_years",
    "payments_applied_during_the_year", "out_of_scope_reason", "year_of_income_being_tested",
]
TEXT_COLUMNS = {"loan_id", "borrower_reference", "year_loan_made", "out_of_scope_reason",
                "year_of_income_being_tested"}
BOOL_COLUMNS = ["written_agreement", "terms_in_place_before_lodgment_day",
                "secured_by_registered_mortgage_over_real_property"]
NUMBER_COLUMNS = ["maximum_term_years", "security_coverage_at_first_made",
                  "interest_rate_for_years_after_year_loan_made",
                  "amalgamated_loan_unpaid_at_end_of_previous_year", "remaining_term_years",
                  "payments_applied_during_the_year"]
YEAR_COLUMNS = ["year_loan_made", "year_of_income_being_tested"]
MONEY_COLUMNS = {"amalgamated_loan_unpaid_at_end_of_previous_year",
                 "payments_applied_during_the_year", "MYR_required", "Shortfall", "Exposure"}

YEAR = "Summary!$B$2"
YEAR_RATE = "Summary!$B$3"
YEAR_START = "Summary!$B$4"


def style(cell, **kw):
    for k, v in kw.items():
        setattr(cell, k, v)
    return cell


def header(ws, row, titles):
    for col, title in enumerate(titles, 1):
        style(ws.cell(row=row, column=col, value=title), **HEAD)


def add_table(ws, name, ref, formulas=None):
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
    first_col, _, last_col, _ = range_boundaries(ref)
    columns = []
    for i, col in enumerate(range(first_col, last_col + 1), 1):
        title = ws.cell(row=1, column=col).value
        formula = (formulas or {}).get(title)
        columns.append(TableColumn(
            id=i, name=title,
            calculatedColumnFormula=None if formula is None else TableFormula(attr_text=formula),
        ))
    table.tableColumns = columns
    ws.add_table(table)


def cell_value(column, raw):
    """A CSV token as the workbook stores it: numbers as numbers, everything else text."""
    if column in TEXT_COLUMNS or raw == "":
        return raw or None
    if column in BOOL_COLUMNS:
        return raw
    try:
        return float(raw)
    except ValueError:
        return raw


def read_sample():
    with SAMPLE.open(newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def read_rates():
    lines = [ln for ln in RATES.read_text(encoding="utf-8").splitlines() if not ln.startswith("#")]
    meta = {}
    for ln in RATES.read_text(encoding="utf-8").splitlines():
        m = re.match(r"#\s*(reviewed_until|reviewed_on):\s*(\S+)", ln)
        if m:
            meta[m.group(1)] = m.group(2)
    return list(csv.DictReader(lines)), meta


def T(column):
    return f"tblLoans[[#This Row],[{column}]]"


def unknown(expr):
    tokens = ",".join(f'"{t}"' for t in sorted(UNKNOWN_TOKENS))
    return f"OR(LOWER(TRIM({expr}&\"\"))={{{tokens}}})"


def tristate(column):
    t = ",".join(f'"{x}"' for x in sorted(_TRUE))
    f = ",".join(f'"{x}"' for x in sorted(_FALSE))
    v = f"LOWER(TRIM({T(column)}&\"\"))"
    return f'=IF({unknown(T(column))},"UNKNOWN",IF(OR({v}={{{t}}}),"PASS",IF(OR({v}={{{f}}}),"FAIL","UNKNOWN")))'


def year_bad(column):
    y = T(column)
    return (
        f"AND(NOT({unknown(y)}),NOT(IFERROR(AND(LEN({y})=7,MID({y},5,1)=\"-\","
        f"ISNUMBER(VALUE(LEFT({y},4))),ISNUMBER(VALUE(RIGHT({y},2))),"
        f"VALUE(RIGHT({y},2))=MOD(VALUE(LEFT({y},4))+1,100)),FALSE)))"
    )


def bool_bad(column):
    v = f"LOWER(TRIM({T(column)}&\"\"))"
    t = ",".join(f'"{x}"' for x in sorted(_TRUE | _FALSE))
    return f"AND(NOT({unknown(T(column))}),NOT(OR({v}={{{t}}})))"


def number_bad(column):
    """Text where a number belongs, or a negative, which the engine refuses for every
    numeric column (parse_money, parse_rate, parse_ratio); a rate above 1 is refused too."""
    x = T(column)
    ceiling = f",{x}>1" if column == "interest_rate_for_years_after_year_loan_made" else ""
    return f"AND(NOT({unknown(x)}),OR(NOT(ISNUMBER({x})),AND(ISNUMBER({x}),OR({x}<0{ceiling}))))"


def formulas():
    a, lo, b, c = "Limb_a_written", "Limb_lodgment", "Limb_b_rate", "Limb_c_term"
    principal = T("amalgamated_loan_unpaid_at_end_of_previous_year")
    payments = T("payments_applied_during_the_year")
    term = T("maximum_term_years")
    year_start = f"VALUE(LEFT({T('year_loan_made')},4))"
    limbs = [T(a), T(lo), T(b), T(c)]
    fails = ",".join(f'{x}="FAIL"' for x in limbs)
    unknowns = ",".join(f'{x}="UNKNOWN"' for x in limbs)
    v = T("MYR_verdict")
    return {
        "Status": (
            f'=IF(NOT({unknown(T("out_of_scope_reason"))}),"SKIPPED",'
            f'IF(IFERROR({year_start}<{FIRST_REVIEWABLE_YEAR.start_year},FALSE),"SKIPPED","REVIEWED"))'
        ),
        "Floor_year": (
            f'=IF({unknown(T("year_of_income_being_tested"))},{T("year_loan_made")}&"",'
            f'{T("year_of_income_being_tested")}&"")'
        ),
        "Floor_rate": (
            f'=IFERROR(INDEX(tblRates[rate],MATCH({T("Floor_year")},tblRates[year_of_income],0)),"")'
        ),
        a: tristate("written_agreement"),
        lo: tristate("terms_in_place_before_lodgment_day"),
        b: (
            f'=IF(NOT(ISNUMBER({T("interest_rate_for_years_after_year_loan_made")})),"UNKNOWN",'
            f'IF({T("Floor_rate")}="","UNKNOWN",'
            f'IF({T("interest_rate_for_years_after_year_loan_made")}>={T("Floor_rate")},"PASS","FAIL")))'
        ),
        "Allowed_term_raw": (
            f'=IF({T("Limb_secured")}="FAIL",7,IF({T("Limb_secured")}="PASS",'
            f'IF(ISNUMBER({T("security_coverage_at_first_made")}),'
            f'IF({T("security_coverage_at_first_made")}>=1.1,25,7),""),""))'
        ),
        "Limb_secured": tristate("secured_by_registered_mortgage_over_real_property"),
        c: (
            f'=IF(NOT(ISNUMBER({term})),"UNKNOWN",IF({T("Allowed_term_raw")}="",'
            f'IF({term}<=7,"PASS","UNKNOWN"),IF({term}<={T("Allowed_term_raw")},"PASS","FAIL")))'
        ),
        "Gate_verdict": (
            f'=IF({T("Status")}="SKIPPED","",IF(OR({fails}),"NOT_COMPLYING",'
            f'IF(OR({unknowns}),"UNKNOWN","COMPLYING")))'
        ),
        "Max_term_allowed": (
            f'=IF({T("Status")}="SKIPPED","",IF({T("Allowed_term_raw")}<>"",{T("Allowed_term_raw")},'
            f'IF(AND(ISNUMBER({term}),{term}<=7),7,"")))'
        ),
        "Term_used": (
            f'=IF({T("Status")}="SKIPPED","",IF(ISNUMBER({T("remaining_term_years")}),'
            f'ROUNDUP({T("remaining_term_years")},0),""))'
        ),
        "MYR_raw": (
            f'=IF(AND(ISNUMBER({principal}),ISNUMBER({YEAR_RATE}),{YEAR_RATE}>0,'
            f'ISNUMBER({T("Term_used")}),{T("Term_used")}>0),'
            f'ROUND({principal}*{YEAR_RATE}/(1-(1/(1+{YEAR_RATE}))^{T("Term_used")}),2),"")'
        ),
        "MYR_verdict": (
            f'=IF({T("Status")}="SKIPPED","",IF(OR({T("Gate_verdict")}<>"COMPLYING",'
            f'{T("year_loan_made")}&""={YEAR},IFERROR({year_start}>{YEAR_START},FALSE),'
            f'AND(ISNUMBER({T("Term_used")}),{T("Term_used")}<=0),'
            f'AND(ISNUMBER({YEAR_RATE}),{YEAR_RATE}<=0)),"REFUSED",'
            f'IF(OR(NOT(ISNUMBER({YEAR_RATE})),NOT(ISNUMBER({principal})),NOT(ISNUMBER({payments})),'
            f'NOT(ISNUMBER({T("Term_used")}))),"UNKNOWN",'
            f'IF({T("MYR_raw")}-ROUND({payments},2)<=0,"MYR_MET","MYR_SHORT"))))'
        ),
        "MYR_reason": (
            f'=IF({v}="REFUSED",IF({T("Gate_verdict")}<>"COMPLYING","s 109N gate is "&{T("Gate_verdict")}&", so s 109E produces no repayment figure",'
            f'IF({T("year_loan_made")}&""={YEAR},"year of income is the year the loan was made (s 109E(1)(a), s 109P)",'
            f'IF(IFERROR({year_start}>{YEAR_START},FALSE),"loan made after the year of income",'
            f'IF(AND(ISNUMBER({T("Term_used")}),{T("Term_used")}<=0),"nil remaining term under s 109E(6)",'
            f'"nil benchmark rate")))),IF({v}="UNKNOWN",'
            f'IF(NOT(ISNUMBER({YEAR_RATE})),"no reviewed benchmark rate for the year of income",'
            f'IF(NOT(ISNUMBER({principal})),"unpaid balance at end of previous year not established",'
            f'IF(NOT(ISNUMBER({payments})),"payments applied not established",'
            f'"remaining term not established"))),""))'
        ),
        "MYR_required": f'=IF(OR({v}="MYR_MET",{v}="MYR_SHORT"),{T("MYR_raw")},"")',
        "Shortfall": (
            f'=IF(OR({v}="MYR_MET",{v}="MYR_SHORT"),MAX(0,ROUND({T("MYR_raw")}-ROUND({payments},2),2)),"")'
        ),
        "Exposure": f'=IF({v}="MYR_SHORT",{T("Shortfall")},"")',
        "Undecided": (
            f'=IF(AND({T("Status")}="REVIEWED",OR({T("Gate_verdict")}="UNKNOWN",{v}="UNKNOWN")),1,0)'
        ),
        # One ISFORMULA per cell: a multi-cell ISFORMULA inside SUMPRODUCT implicitly
        # intersects in a plain formula and a pasted formula slipped through unflagged.
        "Guard": "=IF(OR(" + ",".join(f"_xlfn.ISFORMULA({T(c)})" for c in INPUT_COLUMNS) + "),1,0)",
        "Input_problem": (
            '=TRIM(IF(OR(' + ",".join(year_bad(cn) for cn in YEAR_COLUMNS) + '),"year label ","")'
            '&IF(OR(' + ",".join(bool_bad(cn) for cn in BOOL_COLUMNS) + '),"true/false/unknown ","")'
            '&IF(OR(' + ",".join(number_bad(cn) for cn in NUMBER_COLUMNS) + '),"number ",""))'
        ),
    }


CALC_ORDER = [
    "Status", "Floor_year", "Floor_rate", "Limb_a_written", "Limb_lodgment", "Limb_b_rate",
    "Limb_secured", "Allowed_term_raw", "Limb_c_term", "Gate_verdict", "Max_term_allowed",
    "Term_used", "MYR_raw", "MYR_verdict", "MYR_reason", "MYR_required", "Shortfall",
    "Exposure", "Undecided", "Guard", "Input_problem",
]


def build() -> None:
    rows = read_sample()
    rates, meta = read_rates()
    calc = formulas()
    assert set(calc) == set(CALC_ORDER), set(calc) ^ set(CALC_ORDER)
    wb = Workbook()

    # 1. Start Here
    ws = wb.active
    ws.title = "Start Here"
    style(ws["A1"], value="Division 7A loan review: s 109N terms and s 109E minimum yearly repayment",
          font=Font(bold=True, size=14, color=PURPLE))
    steps = [
        "1. On Register, paste one amalgamated loan per row over the example rows, using the same "
        "columns as the CSV the command line tool reads. Booleans are true, false or unknown; a "
        "blank is unknown. Rates are fractions (8.77 per cent is 0.0877). Years are written 2026-27.",
        "2. On Summary, choose the year of income to review. The s 109N gate reads each loan's "
        "benchmark from the year the loan was made; the repayment uses the chosen year's rate.",
        "3. Read the calculated columns on Register: each s 109N(1) limb, the gate verdict, the "
        "s 109E(6) repayment, the shortfall and the experimental exposure. REFUSED and UNKNOWN are "
        "answers, never a number; the reason column says why.",
        "4. Review Checks: BLOCKED means the register cannot be read (a bad year label, a value "
        "that is not true, false or unknown, text where a number belongs, a formula pasted into an "
        "input, or no reviewed rate for the year). REVIEW means a loan is not on s 109N terms, a "
        "repayment is short, or a row is undecided. PASS means nothing exposed, nothing undecided.",
        "5. A rising benchmark rate raises the minimum yearly repayment on existing complying "
        "loans, not just on new ones: s 109E(6) uses the current year's rate.",
        "6. Experimental review aid. Not a Division 7A determination. The shortfall is not the "
        "dividend (s 109Y, s 109Q and s 109RB are not modelled) and s 109R is not applied.",
    ]
    for i, text in enumerate(steps, 3):
        ws.cell(row=i, column=1, value=text).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110
    style(ws["A10"], value="Overall status", font=Font(bold=True))
    style(ws["A11"], value="='Review Checks'!B10", font=Font(bold=True, size=14), **CALC)
    ws["A13"] = (f"div7a-loan-review workbook, engine {__version__}. Macro-free. "
                 "Not advice; see DISCLAIMER.md in the repository.")

    # 2. Register: inputs then calculated columns in one table
    ws = wb.create_sheet("Register")
    titles = INPUT_COLUMNS + CALC_ORDER
    header(ws, 1, titles)
    for r, row in enumerate(rows, 2):
        for c, column in enumerate(INPUT_COLUMNS, 1):
            cell = ws.cell(row=r, column=c, value=cell_value(column, row.get(column, "")))
            style(cell, **INPUT)
            if column in TEXT_COLUMNS:
                cell.number_format = TEXT
            elif column in MONEY_COLUMNS:
                cell.number_format = MONEY
        for c, column in enumerate(CALC_ORDER, len(INPUT_COLUMNS) + 1):
            cell = style(ws.cell(row=r, column=c, value=calc[column]), **CALC)
            if column in MONEY_COLUMNS:
                cell.number_format = MONEY
    last_row = len(rows) + 1
    add_table(ws, "tblLoans", f"A1:{get_column_letter(len(titles))}{last_row}",
              {k: v[1:] for k, v in calc.items()})
    for column in BOOL_COLUMNS:
        col = get_column_letter(INPUT_COLUMNS.index(column) + 1)
        dv = DataValidation(type="list", formula1='"true,false,unknown"', allow_blank=True)
        dv.add(f"{col}2:{col}{last_row + 500}")
        ws.add_data_validation(dv)
    for c, column in enumerate(titles, 1):
        ws.column_dimensions[get_column_letter(c)].width = 14 if column in MONEY_COLUMNS else 12
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions[get_column_letter(titles.index("MYR_reason") + 1)].width = 40
    ws.column_dimensions[get_column_letter(titles.index("out_of_scope_reason") + 1)].width = 24
    ws.freeze_panes = "B2"

    # 3. Rates: editable, so a reviewed override is a row the accountant adds and cites
    ws = wb.create_sheet("Rates")
    rate_cols = ["year_of_income", "rate", "rba_table", "rba_series", "rba_month", "source", "seen"]
    header(ws, 1, rate_cols)
    for r, row in enumerate(rates, 2):
        for c, column in enumerate(rate_cols, 1):
            value = float(row[column]) if column == "rate" else row[column]
            cell = style(ws.cell(row=r, column=c, value=value), **INPUT)
            if column == "year_of_income":
                cell.number_format = TEXT
    add_table(ws, "tblRates", f"A1:G{len(rates) + 1}")
    ws.column_dimensions["F"].width = 90
    ws["A{}".format(len(rates) + 4)] = (
        "To review a year outside this table, add a row with the May RBA F5 FILRHLBVS figure "
        "and cite where you read it. A blank source is not a review."
    )

    # 4. Summary
    ws = wb.create_sheet("Summary")
    style(ws["A1"], value="Review", **HEAD)
    style(ws["B1"], value="", **HEAD)
    ws["A2"] = "Year of income under review"
    style(ws["B2"], value=DEFAULT_YEAR, number_format=TEXT, protection=Protection(locked=False), **INPUT)
    dv = DataValidation(type="list", formula1=f"=Rates!$A$2:$A${len(rates) + 1}", allow_blank=False)
    dv.add("B2")
    ws.add_data_validation(dv)
    ws["A3"] = "Benchmark interest rate for that year (s 109N(2))"
    style(ws["B3"], value='=IFERROR(INDEX(tblRates[rate],MATCH(B2,tblRates[year_of_income],0)),"")',
          number_format="0.00%", **CALC)
    ws["A4"] = "Year start"
    style(ws["B4"], value='=IFERROR(VALUE(LEFT(B2,4)),"")', **CALC)
    style(ws["A6"], value="Summary (counts are per question, not per row)", **HEAD)
    style(ws["B6"], value="", **HEAD)
    summary = [
        ("rows_reviewed", '=COUNTIF(tblLoans[Status],"REVIEWED")'),
        ("COMPLYING", '=COUNTIF(tblLoans[Gate_verdict],"COMPLYING")'),
        ("NOT_COMPLYING", '=COUNTIF(tblLoans[Gate_verdict],"NOT_COMPLYING")'),
        ("MYR_MET", '=COUNTIF(tblLoans[MYR_verdict],"MYR_MET")'),
        ("MYR_SHORT", '=COUNTIF(tblLoans[MYR_verdict],"MYR_SHORT")'),
        ("UNKNOWN", "=SUM(tblLoans[Undecided])"),
        ("REFUSED", '=COUNTIF(tblLoans[MYR_verdict],"REFUSED")'),
        ("SKIPPED", '=COUNTIF(tblLoans[Status],"SKIPPED")'),
        ("experimental_total_exposure", "=SUM(tblLoans[Exposure])"),
    ]
    for r, (label, formula) in enumerate(summary, 7):
        ws.cell(row=r, column=1, value=label)
        style(ws.cell(row=r, column=2, value=formula), **CALC)
    ws["B15"].number_format = MONEY
    ws["A17"] = "Overall status"
    style(ws["B17"], value="='Review Checks'!B10", font=Font(bold=True), **CALC)
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 18
    ws.protection.sheet = True

    # 5. Review Checks
    ws = wb.create_sheet("Review Checks")
    header(ws, 1, ["Check", "Result", "Count or value", "Example loan"])

    def offender(cond):
        return f"INDEX(tblLoans[loan_id],SUMPRODUCT(MAX({cond}*(ROW(tblLoans[loan_id])-1))))"

    checks = [
        ("Every register value can be read", '=IF(C2=0,"PASS","BLOCKED")',
         '=SUMPRODUCT(--(tblLoans[Input_problem]<>""))',
         '=IF(C2=0,"",' + offender('(tblLoans[Input_problem]<>"")') + ")"),
        ("No formula in an input cell", '=IF(C3=0,"PASS","BLOCKED")', "=SUM(tblLoans[Guard])",
         '=IF(C3=0,"",' + offender("(tblLoans[Guard]=1)") + ")"),
        ("A reviewed benchmark rate exists for the year of income",
         '=IF(ISNUMBER(Summary!B3),"PASS","BLOCKED")', "=Summary!B2", None),
        ("Every reviewed loan is on s 109N terms", '=IF(C5=0,"PASS","REVIEW")',
         '=COUNTIF(tblLoans[Gate_verdict],"NOT_COMPLYING")',
         '=IF(C5=0,"",' + offender('(tblLoans[Gate_verdict]="NOT_COMPLYING")') + ")"),
        ("No minimum yearly repayment is short", '=IF(C6=0,"PASS","REVIEW")',
         '=COUNTIF(tblLoans[MYR_verdict],"MYR_SHORT")',
         '=IF(C6=0,"",' + offender('(tblLoans[MYR_verdict]="MYR_SHORT")') + ")"),
        ("No row is undecided (UNKNOWN)", '=IF(C7=0,"PASS","REVIEW")', "=SUM(tblLoans[Undecided])",
         '=IF(C7=0,"",' + offender("(tblLoans[Undecided]=1)") + ")"),
        ("Rows refused a repayment figure (read the reason; not a breach by itself)",
         '=IF(C8=0,"PASS","NOTE")', '=COUNTIF(tblLoans[MYR_verdict],"REFUSED")',
         '=IF(C8=0,"",' + offender('(tblLoans[MYR_verdict]="REFUSED")') + ")"),
    ]
    for r, (label, result, detail, example) in enumerate(checks, 2):
        ws.cell(row=r, column=1, value=label)
        style(ws.cell(row=r, column=2, value=result), **CALC)
        style(ws.cell(row=r, column=3, value=detail), **CALC)
        cell = style(ws.cell(row=r, column=4), **CALC)
        if example:
            cell.value = ArrayFormula(cell.coordinate, example)
    last = len(checks) + 1
    style(ws["A10"], value="Overall status", font=Font(bold=True))
    style(ws["B10"], value=(f'=IF(COUNTIF(B2:B{last},"BLOCKED")>0,"BLOCKED",'
                            f'IF(COUNTIF(B2:B{last},"REVIEW")>0,"REVIEW","PASS"))'),
          font=Font(bold=True), **CALC)
    style(ws["A12"], value="Caveats the engine attaches to every result", **HEAD)
    caveats = [
        "COMPLYING means the four limbs of s 109N(1) are established on the operator's own facts. "
        "The lodgment day (s 109D(6)) is asserted, not computed. Refinancing reductions to the "
        "maximum term (s 109N(3A) to (3D)) and the other Subdivision D exclusions are not modelled.",
        "s 109R: the workbook does not decide whether the payments applied are genuine repayments. "
        "The figure used is the operator's assertion.",
        "Any shortfall shown is an experimental review aid, not an ATO assessment and not a "
        "s 109E(1) determination. s 109E(2) makes the dividend the shortfall subject to s 109Y, "
        "and s 109E(1)(d) removes it where s 109Q applies. Neither is modelled, nor is s 109RB.",
        "The s 109E(6) amount is rounded to cents half up. The Act prescribes no rounding; this is "
        "the engine's documented choice. A fractional remaining term is rounded up to a whole "
        "number under the closing words of s 109E(6).",
        "A loan year before 1998-99 is skipped: Division 7A reaches a loan made before 4 December "
        "1997 only where its terms were varied on or after that day (s 109D(5)).",
    ]
    for r, text in enumerate(caveats, 13):
        ws.cell(row=r, column=1, value=text).alignment = Alignment(wrap_text=True)
    fills = {"BLOCKED": "F8D7DA", "REVIEW": "FFF3CD", "PASS": "D4EDDA"}
    for target, ref in ((ws, "B2:B10"), (wb["Start Here"], "A11"), (wb["Summary"], "B17")):
        for word, colour in fills.items():
            target.conditional_formatting.add(
                ref, CellIsRule(operator="equal", formula=[f'"{word}"'],
                                fill=PatternFill("solid", fgColor=colour)))
    ws.column_dimensions["A"].width = 100
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.protection.sheet = True

    # 6. Sources & Version
    ws = wb.create_sheet("Sources & Version")
    header(ws, 1, ["Item", "Value"])
    items = [
        ("Workbook engine version", __version__),
        ("Desktop Excel build used to calculate cached values", ""),
        ("Law", "Income Tax Assessment Act 1936 (Cth) Part III Division 7A, C1936A00027, "
                "compilation in force 1 July 2026: s 109D, 109E, 109N, 109P, 109R"),
        ("Law source", "https://www.legislation.gov.au/C1936A00027/latest/text"),
        ("Benchmark rates", "RBA statistical table F5 Indicator Lending Rates, series FILRHLBVS, "
                            "May figure for each year of income (s 109N(2))"),
        ("Rates reviewed until", meta.get("reviewed_until", "")),
        ("Rates reviewed on", meta.get("reviewed_on", "")),
        ("Rates explainer", "https://duguid.com.au/rates/div7a-benchmark-rate/"),
        ("Disclaimer", "Experimental review aid. Not a Division 7A determination, not an ATO "
                       "assessment, and not tax, legal or financial advice."),
    ]
    for r, (k, v) in enumerate(items, 2):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 110

    OUT.parent.mkdir(exist_ok=True)
    wb.save(OUT)


RECALC = r"""
$ErrorActionPreference = 'Stop'
$t = 0; while ((Get-Process EXCEL -ErrorAction SilentlyContinue) -and $t -lt 60) { Start-Sleep 1; $t++ }
if (Get-Process EXCEL -ErrorAction SilentlyContinue) {
  throw 'Excel is already running; refusing to share its COM server.'
}
$xl = New-Object -ComObject Excel.Application
$xl.Visible = $false; $xl.DisplayAlerts = $false; $xl.AutomationSecurity = 1
try {
  $wb = $xl.Workbooks.Open('%s')
  $sources = $wb.Worksheets.Item('Sources & Version')
  $sources.Range('B3').Value2 = 'Excel ' + $xl.Version + ' build ' + $xl.Build
  $xl.CalculateFullRebuild()
  $tries = 0
  while ($xl.CalculationState -ne 0 -and $tries -lt 600) { Start-Sleep -Milliseconds 100; $tries++ }
  $status = $wb.Worksheets.Item('Review Checks').Range('B10').Text
  $wb.Save()
  $wb.Close($false)
  Write-Output ('overall=' + $status)
} finally { $xl.Quit() }
"""


def recalc() -> None:
    script = RECALC % str(OUT).replace("'", "''")
    result = subprocess.run(["powershell", "-NoProfile", "-Command", script],
                            capture_output=True, text=True)
    if result.returncode:
        raise SystemExit(result.stderr or result.stdout)
    print(result.stdout.strip())
    tmp = OUT.with_suffix(".tmp")
    with zipfile.ZipFile(OUT) as src, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "xl/workbook.xml":
                data = re.sub(rb"<x15ac:absPath[^>]*/>", b"", data)
            dst.writestr(item, data)
    tmp.replace(OUT)


if __name__ == "__main__":
    build()
    if "--no-excel" not in sys.argv:
        recalc()
    print(OUT)
