"""Build workbooks/ato-benchmark-compare.xlsx, the accountant-facing Excel front end.

The workbook is a second implementation of the rules in ``atobenchmark`` written in
ordinary worksheet formulas, so an accountant can run the comparison with desktop
Excel and no Python. It is pinned to the engine by ``tests/test_workbook.py``, which
reads the cached values Excel wrote and compares them with the engine's answer for
the bakery example.

Run from the package directory on a machine with desktop Excel:

    uv run --locked --extra dev python tools/build_workbook.py

openpyxl writes the formulas; a single desktop Excel pass over COM recalculates the
workbook so the shipped file carries real cached values, then the build-machine path
Excel stamps into xl/workbook.xml is stripped.
"""

from __future__ import annotations

import csv
import json
import re
import subprocess
import sys
import zipfile
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import Table, TableColumn, TableFormula, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
from atobenchmark import __version__  # noqa: E402
from atobenchmark.mapping import BUCKETS  # noqa: E402

OUT = ROOT / "workbooks" / "ato-benchmark-compare.xlsx"

PURPLE, LAVENDER, GREY = "5C2D91", "F3F1F6", "E2E0DF"
HEAD = dict(font=Font(bold=True, color="FFFFFF"), fill=PatternFill("solid", fgColor=PURPLE))
INPUT = dict(font=Font(color="1F4E9E"), fill=PatternFill("solid", fgColor=GREY))
CALC = dict(fill=PatternFill("solid", fgColor=LAVENDER))
TEXT = "@"
MONEY = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
PCT = "0.00%"

BUCKET_ORDER = list(BUCKETS)  # turnover .. excluded, matches the Calculation rows
EXPENSE_ROWS = "B10:B17"  # cost_of_sales .. other_expense on the Calculation sheet


def style(cell, **kw):
    for k, v in kw.items():
        setattr(cell, k, v)
    return cell


def header(ws, row, titles):
    for col, title in enumerate(titles, 1):
        style(ws.cell(row=row, column=col, value=title), **HEAD)


def read_bakery() -> tuple[list[tuple[str, Decimal]], list[dict[str, str]]]:
    pnl = []
    with (ROOT / "examples" / "bakery-pnl.csv").open(newline="", encoding="utf-8") as fh:
        for row in csv.reader(fh):
            if len(row) > 1 and re.fullmatch(r"-?\d+(\.\d+)?", row[1] or ""):
                pnl.append((row[0], Decimal(row[1])))
    with (ROOT / "examples" / "bakery-mapping.csv").open(newline="", encoding="utf-8") as fh:
        mapping = list(csv.DictReader(fh))
    return pnl, mapping


def read_bands() -> tuple[list[list], dict[str, dict]]:
    rows, sources = [], {}
    for path in sorted((ROOT / "atobenchmark" / "data").glob("benchmarks-*.json")):
        data = json.loads(path.read_text(encoding="utf-8"))
        year, src = data["benchmark_year"], data["source"]
        sources[year] = src
        for bt in data["business_types"]:
            for band in bt["turnover_bands"]:
                cos = band["cost_of_sales_to_turnover"]
                te = band["total_expenses_to_turnover"]
                rows.append(
                    [
                        year,
                        bt["name"],
                        bt["key_ratio"],
                        band["band"],
                        band["label"],
                        float(band["turnover_from"]),
                        "Y" if band["turnover_from_inclusive"] else "N",
                        None if band["turnover_to"] is None else float(band["turnover_to"]),
                        None if cos is None else float(cos["min"]),
                        None if cos is None else float(cos["max"]),
                        None if te is None else float(te["min"]),
                        None if te is None else float(te["max"]),
                        src["retrieved"],
                        src["sha256"],
                    ]
                )
    return rows, sources


def add_table(ws, name, ref, formulas=None):
    """Add a table; ``formulas`` maps column name to a calculated-column formula.

    A calculated column is what makes Excel fill the formula into rows an accountant
    pastes below the example, instead of leaving them blank and BLOCKED.
    """
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
    first_col, _, last_col, _ = range_boundaries(ref)
    columns = []
    for i, col in enumerate(range(first_col, last_col + 1), 1):
        name = ws.cell(row=1, column=col).value
        formula = (formulas or {}).get(name)
        columns.append(
            TableColumn(
                id=i,
                name=name,
                calculatedColumnFormula=None if formula is None else TableFormula(attr_text=formula),
            )
        )
    table.tableColumns = columns
    ws.add_table(table)


def build() -> None:
    pnl, mapping = read_bakery()
    bands, sources = read_bands()
    wb = Workbook()

    # 1. Start Here
    ws = wb.active
    ws.title = "Start Here"
    style(
        ws["A1"],
        value="ATO small business benchmark comparison",
        font=Font(bold=True, size=14, color=PURPLE),
    )
    steps = [
        "1. On P&L Import, paste account names and amounts over the example rows. "
        "Account cells are Text; nothing pasted there is evaluated.",
        "2. On Mapping, give every account a bucket and set Source to reviewed. "
        "A suggested or missing mapping blocks the result.",
        "3. On Calculation, choose the benchmark year and industry. Enter W1 if the activity "
        "statement labour figure is higher than wages. Set B4 to Y if the export shows "
        "expenses as negatives.",
        "4. Read Results and Review Checks. The overall status below is BLOCKED, REVIEW or "
        "PASS; blank or an error is never a pass.",
        "5. Sitting outside a published range is not a finding. The ATO ranges are review "
        "prompts, and the judgement stays with you. Nothing leaves this workbook.",
        "6. Benchmarks and Sources & Version hold the published ranges, their provenance and "
        "the engine version this workbook was checked against.",
    ]
    for i, text in enumerate(steps, 3):
        ws.cell(row=i, column=1, value=text).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110
    style(ws["A10"], value="Overall status", font=Font(bold=True))
    style(ws["A11"], value="='Review Checks'!B13", font=Font(bold=True, size=14), **CALC)
    ws["A13"] = (
        f"ato-benchmark-compare workbook, engine {__version__}. Macro-free. "
        "Not advice; see DISCLAIMER.md in the repository."
    )

    # 2. P&L Import: tblPnl = Account, Amount, Bucket, Guard
    ws = wb.create_sheet("P&L Import")
    header(ws, 1, ["Account", "Amount", "Bucket", "Guard", "Sample"])
    # A fabricated bakery line left in the P&L would count in the ratios; flag the
    # exact account-and-amount pairs the shipped sample carries.
    sample_names = ",".join(f'"{account.strip().lower()}"' for account, _ in pnl)
    sample_amounts = ",".join(str(float(amount)) for _, amount in pnl)
    sample_formula = (
        "=IF(SUMPRODUCT((LOWER(TRIM(tblPnl[[#This Row],[Account]]))={" + sample_names + "})"
        "*(tblPnl[[#This Row],[Amount]]={" + sample_amounts + "}))>0,1,0)"
    )
    bucket_formula = (
        "=IFERROR(INDEX(tblMapping[Bucket],MATCH(TRIM(LOWER(tblPnl[[#This Row],[Account]])),"
        'tblMapping[Key],0)),"")'
    )
    guard_formula = (
        "=IF(OR(_xlfn.ISFORMULA(tblPnl[[#This Row],[Account]]),"
        "_xlfn.ISFORMULA(tblPnl[[#This Row],[Amount]]),"
        'AND(tblPnl[[#This Row],[Account]]<>"",'
        "NOT(ISNUMBER(tblPnl[[#This Row],[Amount]])))),1,0)"
    )
    for r, (account, amount) in enumerate(pnl, 2):
        style(ws.cell(row=r, column=1, value=account), number_format=TEXT, **INPUT)
        style(ws.cell(row=r, column=2, value=float(amount)), number_format=MONEY, **INPUT)
        style(ws.cell(row=r, column=3, value=bucket_formula), **CALC)
        style(ws.cell(row=r, column=4, value=guard_formula), **CALC)
        style(ws.cell(row=r, column=5, value=sample_formula), **CALC)
    add_table(
        ws,
        "tblPnl",
        f"A1:E{len(pnl) + 1}",
        {"Bucket": bucket_formula[1:], "Guard": guard_formula[1:], "Sample": sample_formula[1:]},
    )
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 24

    # 3. Mapping: tblMapping = Account, Bucket, Source, Note, Key
    ws = wb.create_sheet("Mapping")
    header(ws, 1, ["Account", "Bucket", "Source", "Note", "Key"])
    for r, row in enumerate(mapping, 2):
        style(ws.cell(row=r, column=1, value=row["account"]), number_format=TEXT, **INPUT)
        style(ws.cell(row=r, column=2, value=row["bucket"]), **INPUT)
        style(ws.cell(row=r, column=3, value=row["source"]), **INPUT)
        style(ws.cell(row=r, column=4, value=row["note"]), number_format=TEXT, **INPUT)
        key = "=TRIM(LOWER(tblMapping[[#This Row],[Account]]))"
        style(ws.cell(row=r, column=5, value=key), **CALC)
    n = len(mapping) + 1
    add_table(ws, "tblMapping", f"A1:E{n}", {"Key": key[1:]})
    dv = DataValidation(type="list", formula1='"' + ",".join(BUCKET_ORDER) + '"', allow_blank=False)
    dv.add(f"B2:B{n + 200}")
    ws.add_data_validation(dv)
    dv = DataValidation(type="list", formula1='"suggested,reviewed"', allow_blank=False)
    dv.add(f"C2:C{n + 200}")
    ws.add_data_validation(dv)
    for col, width in zip("ABCDE", (34, 24, 12, 60, 34)):
        ws.column_dimensions[col].width = width

    # 4. Benchmarks: tblBands plus the industry list used for validation
    ws = wb.create_sheet("Benchmarks")
    cols = [
        "Year", "Industry", "KeyRatio", "Band", "Label", "From", "FromInclusive", "To",
        "CostOfSalesMin", "CostOfSalesMax", "TotalExpensesMin", "TotalExpensesMax",
        "Retrieved", "Sha256",
    ]
    header(ws, 1, cols)
    for r, row in enumerate(bands, 2):
        for c, value in enumerate(row, 1):
            ws.cell(row=r, column=c, value=value)
    add_table(ws, "tblBands", f"A1:N{len(bands) + 1}")
    industries = sorted({row[1] for row in bands})
    style(ws["P1"], value="Industries", **HEAD)
    for r, name in enumerate(industries, 2):
        ws.cell(row=r, column=16, value=name)
    industry_ref = f"Benchmarks!$P$2:$P${len(industries) + 1}"
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["P"].width = 48
    ws.protection.sheet = True

    # 5. Calculation
    ws = wb.create_sheet("Calculation")
    style(ws["A1"], value="Inputs", **HEAD)
    style(ws["B1"], value="", **HEAD)
    inputs = [
        (2, "Benchmark year", sorted(sources)[-1]),
        (3, "Industry", "Bakeries and hot bread shops"),
        (4, "Expenses shown as negatives? (Y/N)", "N"),
        (5, "Activity statement W1 (blank if not supplied)", None),
    ]
    for r, label, value in inputs:
        ws.cell(row=r, column=1, value=label)
        style(ws.cell(row=r, column=2, value=value), protection=Protection(locked=False), **INPUT)
    ws["B5"].number_format = MONEY
    ws["A6"] = "Expense sign factor"
    ws["B6"] = '=IF(B4="Y",-1,1)'
    years = ",".join(sorted(sources))
    dv = DataValidation(type="list", formula1=f'"{years}"', allow_blank=False)
    dv.add("B2")
    ws.add_data_validation(dv)
    dv = DataValidation(type="list", formula1=f"={industry_ref}", allow_blank=False)
    dv.add("B3")
    ws.add_data_validation(dv)
    dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
    dv.add("B4")
    ws.add_data_validation(dv)

    style(ws["A7"], value="Bucket totals (SUMIFS over tblPnl)", **HEAD)
    style(ws["B7"], value="", **HEAD)
    style(ws["C7"], value="What the ATO does with it", **HEAD)
    for r, bucket in enumerate(BUCKET_ORDER, 8):
        ws.cell(row=r, column=1, value=bucket)
        sign = "*$B$6" if 10 <= r <= 17 else ""
        ws.cell(row=r, column=2, value=f"=SUMIFS(tblPnl[Amount],tblPnl[Bucket],A{r}){sign}")
        ws.cell(row=r, column=3, value=BUCKETS[bucket])

    # First matching band, as the engine does: the minimum matching row index, with
    # non-matching rows pushed to 1E9 so MIN ignores them. 1E9 then means no band.
    band_match = (
        "((tblBands[Year]=B2)*(tblBands[Industry]=B3)"
        '*(((B23>tblBands[From])+((tblBands[FromInclusive]="Y")*(B23=tblBands[From])))>0)'
        '*(((tblBands[To]="")+(B23<=tblBands[To]))>0))'
    )
    band_row_raw = f"=SUMPRODUCT(MIN({band_match}*(ROW(tblBands[Year])-1)+(1-{band_match})*1E9))"
    industry_row = (
        "=SUMPRODUCT(MAX((tblBands[Year]=B2)*(tblBands[Industry]=B3)*(ROW(tblBands[Year])-1)))"
    )
    calc = [
        (20, "Sales of goods and services", "=B8"),
        (21, "Other business income", "=B9"),
        (22, "Total business income", "=B20+B21"),
        (23, "Turnover", "=IF(OR(B20<=0,B20*2<B22),B22,B20)"),
        (24, "Turnover basis",
         '=IF(OR(B20<=0,B20*2<B22),"total business income","sales of goods and services")'),
        (25, "Total expenses reported", f"=SUM({EXPENSE_ROWS})"),
        (26, "Payments to associated persons", "=B14"),
        (27, "Total expenses for ratio", "=B25-B26"),
        (28, "Cost of sales for ratio (excludes wages)", "=B10"),
        (29, "Salary and wages label (wages + cost of sales wages + associates)", "=B12+B11+B26"),
        (30, "Labour base (W1 when it exceeds the label)", '=IF(AND(B5<>"",B5>B29),B5,B29)'),
        (31, "Labour (base less associates plus contractors and commission)", "=B30-B26+B13"),
        (32, "Band lookup (1E9 = no match)", band_row_raw),
        (33, "Band row in tblBands (0 = none)", "=IF(B32>=1E9,0,B32)"),
        (34, "Turnover band", '=IF(B33=0,"none applies",INDEX(tblBands[Label],B33))'),
        (35, "Industry row in tblBands (0 = not published for the year)", industry_row),
        (36, "Published key ratio", '=IF(B35=0,"",INDEX(tblBands[KeyRatio],B35))'),
        (37, "Key ratio used",
         '=IF(AND(B36="cost_of_sales_to_turnover",B28=0),"total_expenses_to_turnover",B36)'),
    ]
    style(ws["A19"], value="Figures", **HEAD)
    style(ws["B19"], value="", **HEAD)
    for r, label, formula in calc:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=formula)
    for r in list(range(8, 19)) + list(range(20, 32)):
        style(ws.cell(row=r, column=2), number_format=MONEY, **CALC)
    for r in (32, 33, 34, 35, 36, 37):
        style(ws.cell(row=r, column=2), **CALC)
    ws.column_dimensions["A"].width = 62
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 70
    ws.protection.sheet = True

    # 6. Results
    ws = wb.create_sheet("Results")
    header(ws, 1, ["Ratio", "Label", "Actual", "Benchmark min", "Benchmark max", "Status", "Key"])
    ratios = [
        ("cost_of_sales_to_turnover", "Cost of sales to turnover", "B28",
         "CostOfSalesMin", "CostOfSalesMax"),
        ("total_expenses_to_turnover", "Total expenses to turnover", "B27",
         "TotalExpensesMin", "TotalExpensesMax"),
        ("labour_to_turnover", "Labour to turnover", "B31", None, None),
        ("rent_to_turnover", "Rent to turnover", "B15", None, None),
        ("motor_vehicle_to_turnover", "Motor vehicle expenses to turnover", "B16", None, None),
    ]
    # ponytail: the published datasets only carry cost of sales and total expenses ranges,
    # so the other three ratios have no benchmark columns; add columns if the ATO adds ranges.
    for r, (key, label, cell, lo, hi) in enumerate(ratios, 2):
        ws.cell(row=r, column=1, value=key)
        ws.cell(row=r, column=2, value=label)
        actual = f'=IFERROR(ROUND(Calculation!{cell}/Calculation!B23,4),"")'
        style(ws.cell(row=r, column=3, value=actual), number_format=PCT, **CALC)
        for col, name in ((4, lo), (5, hi)):
            if name:
                formula = (
                    f'=IF(OR(Calculation!$B$33=0,INDEX(tblBands[{name}],Calculation!$B$33)=""),'
                    f'"",INDEX(tblBands[{name}],Calculation!$B$33))'
                )
            else:
                formula = '=""'
            style(ws.cell(row=r, column=col, value=formula), number_format=PCT, **CALC)
        status = (
            f'=IF(C{r}="","not calculated",IF(Calculation!$B$33=0,"no turnover band applies",'
            f'IF(D{r}="","no benchmark in this dataset",IF(AND(C{r}>=D{r},C{r}<=E{r}),"within",'
            f'IF(C{r}<D{r},"below","above")))))'
        )
        style(ws.cell(row=r, column=6, value=status), **CALC)
        style(ws.cell(row=r, column=7, value=f'=IF(A{r}=Calculation!$B$37,"key","")'), **CALC)
    ws["A8"] = "Turnover"
    style(ws["B8"], value="=Calculation!B23", number_format=MONEY, **CALC)
    ws["A9"] = "Turnover band"
    style(ws["B9"], value="=Calculation!B34", **CALC)
    ws["A10"] = "Benchmark year and industry"
    style(ws["B10"], value='=Calculation!B2&" "&Calculation!B3', **CALC)
    for col, width in zip("ABCDEFG", (28, 34, 12, 16, 16, 28, 6)):
        ws.column_dimensions[col].width = width
    ws.protection.sheet = True

    # 7. Review Checks
    ws = wb.create_sheet("Review Checks")
    header(ws, 1, ["Check", "Result", "Count or value", "Example account"])

    def offender(table, cond):
        """The account on the last row where ``cond`` holds, so the accountant can find it."""
        return (
            f"INDEX({table}[Account],SUMPRODUCT(MAX({cond}*(ROW({table}[Account])-1))))"
        )

    pnl_dup = "(COUNTIF(tblPnl[Account],tblPnl[Account])>1)*(tblPnl[Account]<>\"\")"
    map_dup = "(COUNTIF(tblMapping[Key],tblMapping[Key])>1)*(tblMapping[Key]<>\"\")"
    # ponytail: COUNTIF is case-insensitive like the engine, but it reads ? and * in an
    # account name as wildcards; switch to SUMPRODUCT(--(range=cell)) if that ever bites.
    checks = [
        ("Every P&L account has a bucket", '=IF(C2=0,"PASS","BLOCKED")',
         '=COUNTIFS(tblPnl[Account],"<>",tblPnl[Bucket],"")',
         '=IF(C2=0,"",' + offender("tblPnl", '(tblPnl[Account]<>"")*(tblPnl[Bucket]="")') + ")"),
        ("No mapping is still suggested", '=IF(C3=0,"PASS","BLOCKED")',
         '=COUNTIF(tblMapping[Source],"suggested")',
         '=IF(C3=0,"",' + offender("tblMapping", '(tblMapping[Source]="suggested")') + ")"),
        ("No formula or non-numeric cell in the P&L", '=IF(C4=0,"PASS","BLOCKED")',
         "=SUM(tblPnl[Guard])",
         '=IF(C4=0,"",' + offender("tblPnl", "(tblPnl[Guard]=1)") + ")"),
        ("No P&L account name appears twice", '=IF(C5=0,"PASS","BLOCKED")',
         f"=SUMPRODUCT({pnl_dup})",
         '=IF(C5=0,"",' + offender("tblPnl", pnl_dup) + ")"),
        ("No mapping account appears twice", '=IF(C6=0,"PASS","BLOCKED")',
         f"=SUMPRODUCT({map_dup})",
         '=IF(C6=0,"",' + offender("tblMapping", map_dup) + ")"),
        ("W1 is not negative", '=IF(AND(Calculation!B5<>"",Calculation!B5<0),"BLOCKED","PASS")',
         "=Calculation!B5", None),
        ("Turnover is positive", '=IF(Calculation!B23>0,"PASS","BLOCKED")', "=Calculation!B23",
         None),
        ("Industry is published for the benchmark year", '=IF(Calculation!B35=0,"BLOCKED","PASS")',
         '=Calculation!B2&" "&Calculation!B3', None),
        ("Turnover falls in a published band", '=IF(Calculation!B33=0,"REVIEW","PASS")',
         "=Calculation!B34", None),
        ("Key ratio is within its range", '=IF(C11="within","PASS","REVIEW")',
         '=IFERROR(INDEX(Results!F2:F6,MATCH(Calculation!B37,Results!A2:A6,0)),"not calculated")',
         None),
        ("No fabricated example account from the shipped bakery sample remains in the P&L",
         '=IF(C12=0,"PASS","REVIEW")', "=SUM(tblPnl[Sample])",
         '=IF(C12=0,"",' + offender("tblPnl", "(tblPnl[Sample]=1)") + ")"),
    ]
    for r, (label, result, detail, example) in enumerate(checks, 2):
        ws.cell(row=r, column=1, value=label)
        style(ws.cell(row=r, column=2, value=result), **CALC)
        style(ws.cell(row=r, column=3, value=detail), **CALC)
        # Array formula: a COUNTIF nested inside MAX gets implicit intersection in a
        # plain formula and the example silently names the wrong account.
        cell = ws.cell(row=r, column=4)
        style(cell, **CALC)
        if example:
            cell.value = ArrayFormula(cell.coordinate, example)
    last = len(checks) + 1
    style(ws["A13"], value="Overall status", font=Font(bold=True))
    overall = (
        f'=IF(COUNTIF(B2:B{last},"BLOCKED")>0,"BLOCKED",'
        f'IF(COUNTIF(B2:B{last},"REVIEW")>0,"REVIEW","PASS"))'
    )
    style(ws["B13"], value=overall, font=Font(bold=True), **CALC)
    style(
        ws["A15"],
        value="Notes (the engine prints these as warnings; they do not change the status)",
        **HEAD,
    )
    notes = [
        '=IF(Calculation!B24="total business income","Sales of goods and services are not '
        "positive or are less than half of total business income, so total business income is "
        'used as turnover. That is the ATO rule, but check the mapping.","")',
        '=IF(AND(Calculation!B5<>"",Calculation!B5>Calculation!B29),"Activity statement W1 is '
        'greater than the salary and wages label, so W1 is used in the labour ratio.","")',
        '=IF(AND(Calculation!B5="",Calculation!B11+Calculation!B12>0),"No activity statement W1 '
        "amount was supplied. The ATO uses W1 for the labour ratio when it exceeds the salary "
        'and wages figure.","")',
        '=IF(Calculation!B26=0,"No payments to associated persons were mapped. They are deducted '
        'from total expenses, so a zero here raises the total expenses ratio.","")',
        '=IF(AND(Calculation!B28>0,Calculation!B11=0),"No salary and wages were mapped inside '
        "cost of sales. The ATO excludes wages from the cost of sales ratio, so confirm none are "
        'sitting in those accounts.","")',
        '=IF(MIN(Calculation!B10:B17)<0,"An expense bucket is negative. Check the sign '
        'convention of the export, or set Calculation!B4 to Y.","")',
        '=IF(Calculation!B36="cost_of_sales_to_turnover",IF(Calculation!B28=0,"Cost of sales is '
        "the ATO key range for this industry but none was mapped, so total expenses to turnover "
        'is the key range used here.","The ATO says to use total expenses to turnover as the key '
        "range where cost of sales is only a small amount. Both ranges are reported and that "
        'judgement is yours."),"")',
    ]
    for r, formula in enumerate(notes, 16):
        style(ws.cell(row=r, column=1, value=formula), alignment=Alignment(wrap_text=True), **CALC)
    fills = {
        "BLOCKED": PatternFill("solid", fgColor="F8D7DA"),
        "REVIEW": PatternFill("solid", fgColor="FFF3CD"),
        "PASS": PatternFill("solid", fgColor="D4EDDA"),
    }
    for target, ref in ((ws, "B2:B13"), (wb["Start Here"], "A11")):
        for word, fill in fills.items():
            rule = CellIsRule(operator="equal", formula=[f'"{word}"'], fill=fill)
            target.conditional_formatting.add(ref, rule)
    ws.column_dimensions["A"].width = 100
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 34
    ws.protection.sheet = True

    # 8. Sources & Version
    ws = wb.create_sheet("Sources & Version")
    header(ws, 1, ["Item", "Value"])
    rows = [
        ("Workbook engine version", __version__),
        ("Desktop Excel build used to calculate cached values", ""),
    ]
    for year, src in sorted(sources.items()):
        rows.append((f"{year} publisher", src["publisher"]))
        rows.append((f"{year} dataset", src["dataset"]))
        rows.append((f"{year} resource", src["resource_url"]))
        rows.append((f"{year} retrieved", src["retrieved"]))
        rows.append((f"{year} sha256", src["sha256"]))
        rows.append((f"{year} licence", f'{src["licence"]} {src["licence_url"]}'))
    for r, (k, v) in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 110

    OUT.parent.mkdir(exist_ok=True)
    wb.save(OUT)


RECALC = r"""
$ErrorActionPreference = 'Stop'
if (Get-Process EXCEL -ErrorAction SilentlyContinue) {
  throw 'Excel is already running; refusing to share its COM server.'
}
$xl = New-Object -ComObject Excel.Application
$xl.Visible = $false; $xl.DisplayAlerts = $false; $xl.AutomationSecurity = 1
try {
  $wb = $xl.Workbooks.Open('%s')
  $sources = $wb.Worksheets.Item('Sources & Version')
  $sources.Range('B3').Value2 = 'Excel ' + $xl.Version + ' build ' + $xl.Build
  $xl.CalculateFullRebuild()
  $tries = 0
  while ($xl.CalculationState -ne 0 -and $tries -lt 600) { Start-Sleep -Milliseconds 100; $tries++ }
  $status = $wb.Worksheets.Item('Review Checks').Range('B13').Text
  $wb.Save()
  $wb.Close($false)
  Write-Output ('overall=' + $status)
} finally { $xl.Quit() }
"""


def recalc() -> None:
    script = RECALC % str(OUT).replace("'", "''")
    result = subprocess.run(
        ["powershell", "-NoProfile", "-Command", script], capture_output=True, text=True
    )
    if result.returncode:
        raise SystemExit(result.stderr or result.stdout)
    print(result.stdout.strip())
    # Excel stamps the saving machine's directory into xl/workbook.xml. Strip it.
    tmp = OUT.with_suffix(".tmp")
    with zipfile.ZipFile(OUT) as src, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "xl/workbook.xml":
                data = re.sub(rb"<x15ac:absPath[^>]*/>", b"", data)
            dst.writestr(item, data)
    tmp.replace(OUT)


if __name__ == "__main__":
    build()
    if "--no-excel" not in sys.argv:
        recalc()
    print(OUT)
